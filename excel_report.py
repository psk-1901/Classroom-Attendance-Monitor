"""
Excel report generation module.

This module handles generating professional Excel reports from attendance data
with proper formatting, colors, and calculations.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

from config import config
from database import db
from student import student_manager
from attendance import attendance_manager

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Excel report generation class."""
    
    def __init__(self):
        """Initialize the Excel report generator."""
        self.attendance_dir = config.ATTENDANCE_DIR
        self.attendance_dir.mkdir(exist_ok=True)
        
        # Define styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.present_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        self.absent_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        self.title_font = Font(name='Arial', size=14, bold=True)
        self.subtitle_font = Font(name='Arial', size=11, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
    
    def generate_attendance_report(self, session_id: str, college_name: str = None) -> Path:
        """
        Generate an Excel attendance report for a session.
        
        Args:
            session_id: Attendance session ID
            college_name: College name (optional)
            
        Returns:
            Path to the generated Excel file
        """
        try:
            # Get attendance records
            records = attendance_manager.get_attendance_by_session(session_id)
            
            if not records:
                raise ValueError(f"No attendance records found for session {session_id}")
            
            # Get session info
            session_info = self._get_session_info(session_id)
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Add header information
            row = 1
            
            # College name
            college = college_name or config.EXCEL_TEMPLATE.get('college_name', 'College Name')
            ws.merge_cells(f'A{row}:I{row}')
            ws[f'A{row}'] = college
            ws[f'A{row}'].font = self.title_font
            ws[f'A{row}'].alignment = self.center_alignment
            row += 1
            
            # Title
            ws.merge_cells(f'A{row}:I{row}')
            ws[f'A{row}'] = "ATTENDANCE REPORT"
            ws[f'A{row}'].font = Font(name='Arial', size=16, bold=True)
            ws[f'A{row}'].alignment = self.center_alignment
            row += 2
            
            # Session details
            details = [
                ("Department:", session_info.get('department', 'N/A')),
                ("Section:", session_info.get('section', 'N/A')),
                ("Subject:", session_info.get('subject', 'N/A')),
                ("Faculty:", session_info.get('faculty', 'N/A')),
                ("Date:", session_info.get('date', datetime.now().strftime('%Y-%m-%d'))),
                ("Time:", session_info.get('start_time', 'N/A')),
                ("Classroom:", session_info.get('classroom', 'N/A'))
            ]
            
            for label, value in details:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.subtitle_font
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = self.normal_font
                row += 1
            
            row += 1
            
            # Create headers
            headers = ['Roll No', 'Student Name', 'Department', 'Section', 'Status', 'Time', 'Confidence']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            row += 1
            
            # Add data rows
            present_count = 0
            absent_count = 0
            
            for record in records:
                status = record.get('status', 'absent').upper()
                
                if status == 'PRESENT':
                    present_count += 1
                    fill = self.present_fill
                else:
                    absent_count += 1
                    fill = self.absent_fill
                
                data = [
                    record.get('roll_number', 'N/A'),
                    record.get('name', 'N/A'),
                    record.get('department', 'N/A'),
                    record.get('section', 'N/A'),
                    status,
                    record.get('time_in', '--'),
                    f"{record.get('confidence', 0)*100:.1f}%" if status == 'PRESENT' else '--'
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = self.normal_font
                    cell.alignment = self.center_alignment if col_idx == 5 else self.left_alignment
                    cell.border = self.thin_border
                    
                    if col_idx == 5:  # Status column
                        cell.fill = fill
                
                row += 1
            
            # Add summary
            row += 1
            total = present_count + absent_count
            
            ws[f'A{row}'] = "SUMMARY"
            ws[f'A{row}'].font = self.subtitle_font
            row += 1
            
            summary_data = [
                ("Total Students:", total),
                ("Present:", present_count),
                ("Absent:", absent_count),
                ("Attendance Percentage:", f"{(present_count/total*100):.1f}%" if total > 0 else "0.0%")
            ]
            
            for label, value in summary_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.subtitle_font
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = Font(name='Arial', size=10, bold=True) if 'Percentage' in label else self.normal_font
                row += 1
            
            # Auto adjust column widths
            for col in range(1, 8):
                max_length = 0
                column = get_column_letter(col)
                for row_cell in range(1, row):
                    cell = ws[f'{column}{row_cell}']
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2
            
            # Freeze panes
            ws.freeze_panes = 'A8'
            
            # Apply conditional formatting for status column
            self._apply_conditional_formatting(ws)
            
            # Save file
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"Attendance_{session_id}_{date_str}.xlsx"
            filepath = self.attendance_dir / filename
            
            # Create date subdirectory
            date_dir = self.attendance_dir / date_str
            date_dir.mkdir(exist_ok=True)
            
            final_path = date_dir / filename
            wb.save(str(final_path))
            
            logger.info(f"Attendance report generated: {final_path}")
            return final_path
            
        except Exception as e:
            logger.error(f"Error generating attendance report: {e}")
            raise
    
    def _apply_conditional_formatting(self, worksheet):
        """
        Apply conditional formatting to the status column.
        
        Args:
            worksheet: The worksheet to apply formatting to
        """
        try:
            # Define colors
            green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            # Apply formatting to each cell in status column
            # This is a simpler approach that works with all openpyxl versions
            for row in worksheet.iter_rows(min_row=8, min_col=5, max_col=5):
                for cell in row:
                    if cell.value == 'PRESENT':
                        cell.fill = green_fill
                    elif cell.value == 'ABSENT':
                        cell.fill = red_fill
                        
        except Exception as e:
            logger.warning(f"Could not apply conditional formatting: {e}")
    
    def generate_department_report(self, department: str, date: str = None) -> Path:
        """
        Generate attendance report for a department.
        
        Args:
            department: Department name
            date: Date string (YYYY-MM-DD), defaults to today
            
        Returns:
            Path to the generated Excel file
        """
        try:
            if date is None:
                date = datetime.now().strftime('%Y-%m-%d')
            
            # Get attendance records for the department
            records = attendance_manager.get_attendance_by_date(date)
            dept_records = [r for r in records if r.get('department') == department]
            
            if not dept_records:
                raise ValueError(f"No attendance records found for {department} on {date}")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{department} Attendance"
            
            # Add header
            ws['A1'] = f"{department} Department Attendance Report"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:G1')
            
            ws['A2'] = f"Date: {date}"
            ws['A2'].font = self.subtitle_font
            
            # Add headers
            headers = ['Roll No', 'Name', 'Section', 'Subject', 'Status', 'Time', 'Faculty']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Add data
            row = 5
            present = 0
            absent = 0
            
            for record in dept_records:
                status = record.get('status', 'absent').upper()
                
                if status == 'PRESENT':
                    present += 1
                    fill = self.present_fill
                else:
                    absent += 1
                    fill = self.absent_fill
                
                data = [
                    record.get('roll_number', 'N/A'),
                    record.get('name', 'N/A'),
                    record.get('section', 'N/A'),
                    record.get('subject', 'N/A'),
                    status,
                    record.get('time_in', '--'),
                    record.get('faculty', 'N/A')
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = self.normal_font
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border
                    
                    if col_idx == 5:
                        cell.fill = fill
                
                row += 1
            
            # Add summary
            total = present + absent
            ws['A' + str(row+2)] = "Summary"
            ws['A' + str(row+2)].font = self.subtitle_font
            
            summary = [
                (f"Total {department} Students:", total),
                ("Present:", present),
                ("Absent:", absent),
                ("Attendance Percentage:", f"{(present/total*100):.1f}%" if total > 0 else "0.0%")
            ]
            
            for i, (label, value) in enumerate(summary):
                ws[f'A{row+3+i}'] = label
                ws[f'B{row+3+i}'] = value
            
            # Adjust column widths
            for col in range(1, 8):
                column = get_column_letter(col)
                ws.column_dimensions[column].width = 15
            
            # Save file
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"{department}_Attendance_{date_str}.xlsx"
            
            dept_dir = self.attendance_dir / 'department_reports'
            dept_dir.mkdir(exist_ok=True)
            filepath = dept_dir / filename
            
            wb.save(str(filepath))
            logger.info(f"Department report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating department report: {e}")
            raise
    
    def generate_student_report(self, roll_number: str) -> Path:
        """
        Generate attendance report for a specific student.
        
        Args:
            roll_number: Student's roll number
            
        Returns:
            Path to the generated Excel file
        """
        try:
            # Get student info
            student = student_manager.get_student(roll_number)
            if not student:
                raise ValueError(f"Student {roll_number} not found")
            
            # Get attendance records
            records = attendance_manager.get_attendance_by_student(roll_number)
            
            if not records:
                raise ValueError(f"No attendance records found for {student['name']}")
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{student['name']} Attendance"
            
            # Add header
            ws['A1'] = f"Student Attendance Report - {student['name']}"
            ws['A1'].font = self.title_font
            ws.merge_cells('A1:G1')
            
            ws['A2'] = f"Roll Number: {roll_number}"
            ws['B2'] = f"Department: {student.get('department', 'N/A')}"
            ws['D2'] = f"Section: {student.get('section', 'N/A')}"
            
            # Add headers
            headers = ['Date', 'Subject', 'Faculty', 'Status', 'Time', 'Classroom', 'Session ID']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.thin_border
            
            # Add data
            row = 5
            present = 0
            absent = 0
            
            for record in records:
                status = record.get('status', 'absent').upper()
                
                if status == 'PRESENT':
                    present += 1
                    fill = self.present_fill
                else:
                    absent += 1
                    fill = self.absent_fill
                
                data = [
                    record.get('date', 'N/A'),
                    record.get('subject', 'N/A'),
                    record.get('faculty', 'N/A'),
                    status,
                    record.get('time_in', '--'),
                    record.get('classroom', 'N/A'),
                    record.get('session_id', 'N/A')
                ]
                
                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.font = self.normal_font
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border
                    
                    if col_idx == 4:
                        cell.fill = fill
                
                row += 1
            
            # Add summary
            total = present + absent
            ws['A' + str(row+2)] = "Summary"
            ws['A' + str(row+2)].font = self.subtitle_font
            
            summary = [
                (f"Total Sessions:", total),
                ("Present:", present),
                ("Absent:", absent),
                ("Attendance Percentage:", f"{(present/total*100):.1f}%" if total > 0 else "0.0%")
            ]
            
            for i, (label, value) in enumerate(summary):
                ws[f'A{row+3+i}'] = label
                ws[f'B{row+3+i}'] = value
            
            # Adjust column widths
            for col in range(1, 8):
                column = get_column_letter(col)
                ws.column_dimensions[column].width = 15
            
            # Save file
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"{roll_number}_{student['name']}_Attendance_{date_str}.xlsx"
            
            student_dir = self.attendance_dir / 'student_reports'
            student_dir.mkdir(exist_ok=True)
            filepath = student_dir / filename
            
            wb.save(str(filepath))
            logger.info(f"Student report generated: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating student report: {e}")
            raise
    
    def _get_session_info(self, session_id: str) -> Dict[str, Any]:
        """
        Get session information from database.
        
        Args:
            session_id: Session ID
            
        Returns:
            Session information dictionary
        """
        try:
            # Try to get from attendance records
            records = attendance_manager.get_attendance_by_session(session_id)
            if records and len(records) > 0:
                record = records[0]
                return {
                    'subject': record.get('subject', 'N/A'),
                    'faculty': record.get('faculty', 'N/A'),
                    'department': record.get('department', 'N/A'),
                    'section': record.get('section', 'N/A'),
                    'date': record.get('date', 'N/A'),
                    'start_time': record.get('time_in', 'N/A'),
                    'classroom': record.get('classroom', 'N/A')
                }
            
            # Try from logs
            logs = db.get_collection(config.COLLECTIONS['logs'])
            session_log = logs.find_one({
                'event_type': 'attendance_session_start',
                'session_id': session_id
            })
            
            if session_log and 'details' in session_log:
                details = session_log['details']
                return {
                    'subject': details.get('subject', 'N/A'),
                    'faculty': details.get('faculty', 'N/A'),
                    'department': details.get('department', 'N/A'),
                    'section': details.get('section', 'N/A'),
                    'date': details.get('date', 'N/A'),
                    'start_time': details.get('start_time', 'N/A'),
                    'classroom': details.get('classroom', 'N/A')
                }
            
            return {}
            
        except Exception as e:
            logger.error(f"Error getting session info: {e}")
            return {}


# Create a singleton instance
excel_report = ExcelReportGenerator()


if __name__ == "__main__":
    # Test report generation
    print("Excel Report Generation")
    print("=" * 50)
    
    # Generate a sample report (requires actual session ID)
    session_id = input("Enter session ID to generate report (or press Enter to skip): ")
    if session_id:
        try:
            filepath = excel_report.generate_attendance_report(session_id)
            print(f"Report generated: {filepath}")
        except Exception as e:
            print(f"Error: {e}")